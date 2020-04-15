from flights import flightsFromFile
from convert import n_to_icao, icao_to_n
from scrape import scrape_individual, print_dict
from openpyxl.workbook import Workbook

"""
For a list of flights, check the faa registration associated with the icao address
"""

allFlights = flightsFromFile('../data/flight_lists/2019-11_2020-03.json')
icaos = set()
for f in allFlights.elements:
    icaos.add(f.icao)
icaos=list(icaos)
l = len(icaos)
print(str(l)+' distincts icaos found')
found = list()
nnumbers = list()
total = 60

count = 0
for i in range(len(icaos)):
    nnumber = icao_to_n(icaos[i])
    if nnumber is None or len(nnumber)!=6:
        continue
    reg = scrape_individual(nnumber)
    if 'Registration' in reg and 'Reserved' in reg['Registration']:
        if 'Type Reservation' in reg and reg['Type Reservation']=='No Fee':
            if 'Reserving Party Name' in reg and reg['Reserving Party Name']=='SBS PROGRAM OFFICE':
                found.append(icaos[i])
                nnumbers.append(nnumber)
                print_dict(reg)
                print()
                count += 1

    #perc = int(total * i / l)
    #print('|'+'█'*(perc)+' '*(total-perc)+'|'+' '+str(i+1)+'/'+str(l)+' '*4+nnumber+' '*4+found,end='\n')

#print('|'+'█'*(total)+'|'+' '+str(i+1)+'/'+str(l)+' '*5+'Done!',end='\r')
print('Found: '+str(found))
print(nnumbers)
print(count)

callsigns = dict()
airports = dict()
for i in found:
    callsigns[i] = list()

for f in allFlights.elements:
    if f.icao in callsigns:
        callsigns[f.icao].append(f.callsign)

        for a in [f.departure.airport, f.arrival.airport]:
            if a not in airports:
                airports[a] = set()
            airports[a].add(f.icao)

#print_dict(callsigns)


wb = Workbook()
ws = wb.active

i = 1

for e in sorted(callsigns):
    ws.cell(row=i, column=1).value = e
    ws.cell(row=i, column=2).value = icao_to_n(e)
    j = 3
    for g in callsigns[e]:
        ws.cell(row=i, column=j).value = g
        j+=1
    i+=1

for e in sorted(airports):
    ws.cell(row=i, column=1).value = e
    j = 2
    for g in airports[e]:
        ws.cell(row=i, column=j).value = g
        j+=1
    i+=1

wb.save('../data/sheets/tmp.xlsx')

#print_dict(airports)

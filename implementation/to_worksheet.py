from flights import Flights
import sys
import re
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook

from os import listdir
from os.path import isfile, join

"""
Export monthly data from ../data/flight_lists/ in Flight structures to Workbook
"""

flightslistpath = '../data/flight_lists/'
sheets_directory = '../data/sheets/'
extension = '.xlsx'

def addSheet(wb, filename):
    month = filename.split('.')[0]

    allFlights = Flights()
    allFlights.from_file(flightslistpath+filename)

    ws = wb.create_sheet(month)
    ws.title = month

    row0 = 1
    row_title = row0
    row_desc = row_title + 1
    row_unit = row_desc + 1
    row_data = row_unit + 1

    col0 = 1
    col_icao = col0
    col_callsign = col_icao + 1
    col_nnumber = col_callsign + 1

    col_dep_ap = col_nnumber + 1
    col_dep_time = col_dep_ap + 1
    col_dep_ap_pos_lat = col_dep_time + 1
    col_dep_ap_pos_lon = col_dep_ap_pos_lat + 1
    col_dep_ap_pos_alt = col_dep_ap_pos_lon + 1
    col_dep_ac_pos_lat = col_dep_ap_pos_alt + 1
    col_dep_ac_pos_lon = col_dep_ac_pos_lat + 1
    col_dep_ac_pos_alt = col_dep_ac_pos_lon + 1

    col_arr_ap = col_dep_ac_pos_alt + 1
    col_arr_time = col_arr_ap + 1
    col_arr_ap_pos_lat = col_arr_time + 1
    col_arr_ap_pos_lon = col_arr_ap_pos_lat + 1
    col_arr_ap_pos_alt = col_arr_ap_pos_lon + 1
    col_arr_ac_pos_lat = col_arr_ap_pos_alt + 1
    col_arr_ac_pos_lon = col_arr_ac_pos_lat + 1
    col_arr_ac_pos_alt = col_arr_ac_pos_lon + 1

    col_last = col_arr_ac_pos_alt + 1

    ws.merge_cells(start_row=row_title, end_row=row_title, start_column=col_icao, end_column=col_nnumber)
    ws.cell(row=row_title, column=col_icao).value = 'Identification Numbers'
    ws.cell(row=row_title, column=col_icao).alignment = Alignment(horizontal='center')
    ws.cell(row=row_desc,column=col_icao).value = 'ICAO24'
    ws.cell(row=row_desc,column=col_callsign).value = 'Callsign'
    ws.cell(row=row_desc,column=col_nnumber).value = 'N-Number'

    ws.merge_cells(start_row=row_title, end_row=row_title, start_column=col_dep_ap, end_column=col_dep_ac_pos_alt)
    ws.cell(row=row_title, column=col_dep_ap).value = "Departure"
    ws.cell(row=row_title, column=col_dep_ap).alignment = Alignment(horizontal='center')
    ws.cell(row=row_desc,column=col_dep_ap).value = 'Departure Airport'
    ws.cell(row=row_unit,column=col_dep_ap).value = 'ICAO number'
    ws.cell(row=row_desc,column=col_dep_time).value = 'Departure Time'
    ws.cell(row=row_unit,column=col_dep_time).value = 'YYYY-MM-DD HH:MM:SS'

    ws.merge_cells(start_row=row_desc, end_row=row_desc, start_column=col_dep_ap_pos_lat, end_column=col_dep_ap_pos_alt)
    ws.cell(row=row_desc,column=col_dep_ap_pos_lat).value = 'Departure Airport Positon'
    ws.cell(row=row_desc,column=col_dep_ap_pos_lat).alignment = Alignment(horizontal='center')
    ws.cell(row=row_unit,column=col_dep_ap_pos_lat).value = 'Latitude'
    ws.cell(row=row_unit,column=col_dep_ap_pos_lon).value = 'Longitude'
    ws.cell(row=row_unit,column=col_dep_ap_pos_alt).value = 'Altitude'

    ws.merge_cells(start_row=row_desc, end_row=row_desc, start_column=col_dep_ac_pos_lat, end_column=col_dep_ac_pos_alt)
    ws.cell(row=row_desc,column=col_dep_ac_pos_lat).value = 'First Aircraft Positon'
    ws.cell(row=row_desc,column=col_dep_ac_pos_lat).alignment = Alignment(horizontal='center')
    ws.cell(row=row_unit,column=col_dep_ac_pos_lat).value = 'Latitude'
    ws.cell(row=row_unit,column=col_dep_ac_pos_lon).value = 'Longitude'
    ws.cell(row=row_unit,column=col_dep_ac_pos_alt).value = 'Altitude'

    ws.merge_cells(start_row=row_title, end_row=row_title, start_column=col_arr_ap, end_column=col_arr_ac_pos_alt)
    ws.cell(row=row_title, column=col_arr_ap).value = "Arrival"
    ws.cell(row=row_title, column=col_arr_ap).alignment = Alignment(horizontal='center')
    ws.cell(row=row_desc,column=col_arr_ap).value = 'Arrival Airport'
    ws.cell(row=row_unit,column=col_arr_ap).value = 'ICAO number'
    ws.cell(row=row_desc,column=col_arr_time).value = 'Departure Time'
    ws.cell(row=row_unit,column=col_arr_time).value = 'YYYY-MM-DD HH:MM:SS'

    ws.merge_cells(start_row=row_desc, end_row=row_desc, start_column=col_arr_ap_pos_lat, end_column=col_arr_ap_pos_alt)
    ws.cell(row=row_desc,column=col_arr_ap_pos_lat).value = 'Arrival Airport Positon'
    ws.cell(row=row_desc,column=col_arr_ap_pos_lat).alignment = Alignment(horizontal='center')
    ws.cell(row=row_unit,column=col_arr_ap_pos_lat).value = 'Latitude'
    ws.cell(row=row_unit,column=col_arr_ap_pos_lon).value = 'Longitude'
    ws.cell(row=row_unit,column=col_arr_ap_pos_alt).value = 'Altitude'

    ws.merge_cells(start_row=row_desc, end_row=row_desc, start_column=col_arr_ac_pos_lat, end_column=col_arr_ac_pos_alt)
    ws.cell(row=row_desc,column=col_arr_ac_pos_lat).value = 'Last Aircraft Positon'
    ws.cell(row=row_desc,column=col_arr_ac_pos_lat).alignment = Alignment(horizontal='center')
    ws.cell(row=row_unit,column=col_arr_ac_pos_lat).value = 'Latitude'
    ws.cell(row=row_unit,column=col_arr_ac_pos_lon).value = 'Longitude'
    ws.cell(row=row_unit,column=col_arr_ac_pos_alt).value = 'Altitude'

    for i in range(col0, col_last):
        ws.column_dimensions[chr(ord('A')-col0+i)].width = 15
    ws.column_dimensions[chr(ord('A')-col0+col_dep_time)].width = 19
    ws.column_dimensions[chr(ord('A')-col0+col_arr_time)].width = 19

    for f in allFlights.elements:
        ws.cell(row=row_data, column=col_icao).value = f.icao.strip()
        ws.cell(row=row_data, column=col_callsign).value = f.callsign.strip()
        if f.nnumber is not None:
        #if 'None' not in f.nnumber:
            ws.cell(row=row_data, column=col_nnumber).value = f.nnumber

        if f.departure.airport is not None:
        #if 'None' not in f.departure.airport:
            ws.cell(row=row_data, column=col_dep_ap).value = f.departure.airport
        ws.cell(row=row_data, column=col_dep_time).value = str(f.departure.time)
        if f.departure.airport_position is not None:
            ws.cell(row=row_data, column=col_dep_ap_pos_lat).value = f.departure.airport_position.latitude
            ws.cell(row=row_data, column=col_dep_ap_pos_lon).value = f.departure.airport_position.longitude
            ws.cell(row=row_data, column=col_dep_ap_pos_alt).value = f.departure.airport_position.altitude
        if f.departure.aircraft_position is not None:
            ws.cell(row=row_data, column=col_dep_ac_pos_lat).value = f.departure.aircraft_position.latitude
            ws.cell(row=row_data, column=col_dep_ac_pos_lon).value = f.departure.aircraft_position.longitude
            ws.cell(row=row_data, column=col_dep_ac_pos_alt).value = f.departure.aircraft_position.altitude


        if f.arrival.airport is not None:
        #if 'None' not in f.arrival.airport:
            ws.cell(row=row_data, column=col_arr_ap).value = f.arrival.airport
        ws.cell(row=row_data, column=col_arr_time).value = str(f.arrival.time)
        if f.arrival.airport_position is not None:
            ws.cell(row=row_data, column=col_arr_ap_pos_lat).value = f.arrival.airport_position.latitude
            ws.cell(row=row_data, column=col_arr_ap_pos_lon).value = f.arrival.airport_position.longitude
            ws.cell(row=row_data, column=col_arr_ap_pos_alt).value = f.arrival.airport_position.altitude
        if f.arrival.aircraft_position is not None:
            ws.cell(row=row_data, column=col_arr_ac_pos_lat).value = f.arrival.aircraft_position.latitude
            ws.cell(row=row_data, column=col_arr_ac_pos_lon).value = f.arrival.aircraft_position.longitude
            ws.cell(row=row_data, column=col_arr_ac_pos_alt).value = f.arrival.aircraft_position.altitude

        row_data += 1
    return wb

def newWorkbook(filename):

    files = [f for f in listdir(flightslistpath) if isfile(join(flightslistpath, f))]

    wb = Workbook()
    wb.remove(wb.active)

    for f in files:
        wb = addSheet(wb, f)

    wb.save(sheets_directory+filename+extension)

def expandWorkbook(filename):
    fullname = sheets_directory+filename+extension
    wb = load_workbook(fullname)

    files = [f for f in listdir(flightslistpath) if isfile(join(flightslistpath, f))]
    sheetnames = wb.sheetnames

    new_files = list()
    for f in files:
        if f.split('.')[0] not in sheetnames:
            new_files.append(f)
    
    if len(new_files)==0:
        print("No update to be made to the given workbook")
        sys.exit()

    rex = re.compile("^[0-9]{4}[-][0-9]{2}[_][0-9]{4}[-][0-9]{2}$") # NNNN-NN_NNNN-NN
    allflightsname=""
    for n in sheetnames:
        if rex.match(n):
            wb.remove(wb[n])
            allflightsname=n

    for f in new_files:
        if rex.match(f.split('.')[0]):
            allflightsname = f.split('.')[0]
        wb = addSheet(wb, f)

    wb._sheets.sort(key=lambda ws: ws.title)
    
    info_index = -1
    index_allflights = -1
    for i in range(len(wb._sheets)):
        if wb._sheets[i].title == 'Info':
            info_index=i
        elif wb._sheets[i].title == allflightsname:
            index_allflights=i

    order = [i for i in range(len(wb._sheets))]
    if index_allflights != -1:
            order.insert(0, order.pop(index_allflights))
    else:
        print('List of all flights not found!')

    if info_index != -1:
        order.insert(0, order.pop(info_index))
    else:
        print('Info sheet not found!')

    wb._sheets =[wb._sheets[i] for i in order]
    wb.active=0

    wb.save(sheets_directory+filename+"_new"+extension)

if __name__ == "__main__":
    #newWorkbook('flights')
    expandWorkbook('DCM_FFL_flights')
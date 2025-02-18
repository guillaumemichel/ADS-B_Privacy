from flights import flightsFromFile
from convert import icao_to_n
from datetime import datetime
from openpyxl.workbook import Workbook

"""
Various scripts to make statistics about the recoreded flights
"""

allFlights = flightsFromFile('../data/flight_lists/2019-06_2020-04.json')

"""
date0 = datetime.strptime('2019-12-30', '%Y-%m-%d')
date1 = datetime.strptime('2020-01-02', '%Y-%m-%d')
icao = 'ad8471'

l = list()
for f in allFlights.elements:
    if f.icao == icao:# and date0 < datetime.strptime(f.departure.time, '%Y-%m-%d %H:%M:%S') < date1:
        print(f.departure.time)

for e in l:
    print(e)
"""

"""
# Print frequencies of callsigns for icaos that were not observed before 2020-01-01
date = datetime.strptime('2020-01-01', '%Y-%m-%d')

icaos = set()
for f in allFlights.elements:
    icaos.add(f.icao)

for f in allFlights.elements:
    if f.icao in icaos and datetime.strptime(f.departure.time, '%Y-%m-%d %H:%M:%S') < date:
        icaos.remove(f.icao)

d = dict()
for f in allFlights.elements:
    if f.icao in icaos:
        if f.icao not in d:
            d[f.icao] = dict()

        if f.callsign not in d[f.icao]:
            d[f.icao][f.callsign] = 0

        d[f.icao][f.callsign]+=1

for e in d:
    print(e+' ',d[e])
"""

"""
# All flights that icao did not appear before a certain date 
date = datetime.strptime('2020-03-01', '%Y-%m-%d')

icaos = set()
for f in allFlights.elements:
    if datetime.strptime(f.departure.time, '%Y-%m-%d %H:%M:%S') > date:
        icaos.add(f.icao)

for f in allFlights.elements:
    if f.icao in icaos and datetime.strptime(f.departure.time, '%Y-%m-%d %H:%M:%S') < date:
        icaos.remove(f.icao)

print(len(icaos))
print(icaos)
"""


"""
# All flights that icao did not appear after 2020-01-01 
date = datetime.strptime('2020-01-01', '%Y-%m-%d')

icaos = set()
for f in allFlights.elements:
    icaos.add(f.icao)

for f in allFlights.elements:
    if f.icao in icaos and datetime.strptime(f.departure.time, '%Y-%m-%d %H:%M:%S') > date:
        icaos.remove(f.icao)

print(icaos)
"""


"""
# For each icao print all callsigns it is associated with and frequencies
d = dict()
for f in allFlights.elements:
    if 'FFL' not in f.callsign and 'DCM' not in f.callsign:
        continue
    if f.icao not in d:
        d[f.icao] = dict()

    if f.callsign not in d[f.icao]:
        d[f.icao][f.callsign] = 0

    d[f.icao][f.callsign]+=1

for e in d:
    if len(d[e])>2:
        print(e+' ',d[e])

print(len(d))
"""
"""
# For each icao print all N-Numbers it is associated with
d = dict()
for f in allFlights.elements:
    if 'FFL' not in f.callsign and 'DCM' not in f.callsign:
        continue

    if f.callsign[0] == 'N':
        if f.icao not in d:
            d[f.icao] = dict()

        if f.callsign not in d[f.icao]:
            d[f.icao][f.callsign] = 0

        d[f.icao][f.callsign]+=1
    if f.nnumber is not None:
        if f.icao not in d:
            d[f.icao] = dict()

        if f.callsign not in d[f.icao]:
            d[f.icao][f.nnumber] = 0

        d[f.icao][f.nnumber]+=1


for e in d:
    if len(d[e])>1:
        print(e+' ',d[e])
"""
"""
# For each callsign print all icaos it is associated with and frequencies
d = dict()
for f in allFlights.elements:
    if 'FFL' not in f.callsign and 'DCM' not in f.callsign:
        continue
    if f.callsign not in d:
        d[f.callsign] = dict()

    if f.icao not in d[f.callsign]:
        d[f.callsign][f.icao] = 0

    d[f.callsign][f.icao]+=1

for e in sorted(d):
    if len(d[e])>1:
        print(e+' ',d[e])
"""

"""
# For each icao print all FFL callsigns it is associated with and frequencies
d = dict()
for f in allFlights.elements:
    if f.icao not in d:
        d[f.icao] = dict()

    if f.callsign not in d[f.icao]:
        d[f.icao][f.callsign] = 0

    d[f.icao][f.callsign]+=1

for e in d:
    c = 0
    for f in d[e]:
        if 'FFL' in f:
            c+=1
    if c>2:
        print(e+' ',d[e])

print(len(d))
"""

"""
# For each flight using N-Number, check if it correspond with its ICAO's conversion
d = dict()
for f in allFlights.elements:
    #if f.nnumber[0]=='N' and f.nnumber != icao_to_n(f.icao):
    if f.callsign is not None and len(f.callsign)>0 and f.callsign[0] == 'N' and f.callsign != icao_to_n(f.icao):
        if f.icao not in d:
            d[f.icao] = dict()

        if f.callsign not in d[f.icao]:
            d[f.icao][f.callsign] = 0

        d[f.icao][f.callsign]+=1

print(d)

wb = Workbook()
ws = wb.active

i = 1
for e in sorted(d):
    if len(d[e])>1:
        ws.cell(row=i, column=1).value = e
        ws.cell(row=i, column=2).value = icao_to_n(e)
        j = 3
        for g in d[e]:
            ws.cell(row=i, column=j).value = g
            ws.cell(row=i+1, column=j).value = d[e][g]
            j+=1

        i+=2

wb.save('../data/sheets/tmp.xlsx')

"""

"""
# Write data (map(key0, map(key1, value))) to temporary workbook (to facilitate export)
wb = Workbook()
ws = wb.active

i = 1
for e in sorted(d):
    if len(d[e])>1:
        ws.cell(row=i, column=1).value = e
        j = 2
        for g in d[e]:
            ws.cell(row=i, column=j).value = g
            ws.cell(row=i+1, column=j).value = d[e][g]
            j+=1

        i+=2

wb.save('../data/sheets/tmp.xlsx')
"""